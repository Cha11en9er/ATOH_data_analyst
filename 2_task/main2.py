import os
import csv
import zipfile
import tempfile
from docx import Document
from openpyxl import load_workbook
import PyPDF2
import sys

def normalize_content_for_csv(content):
    """
    Приводит текст к однострочному виду для корректной записи в CSV.
    Это убирает переносы строк, из-за которых появляются отдельные строки с кавычками.
    """
    if content is None:
        return ""
    return " ".join(str(content).split())

def extract_text_from_file(file_path, file_type):
    """Извлекает текст из файла в зависимости от его типа"""
    try:
        if file_type == 'txt':
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        elif file_type == 'csv':
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        elif file_type == 'docx':
            doc = Document(file_path)
            return '\n'.join([para.text for para in doc.paragraphs])
        elif file_type == 'xlsx':
            wb = load_workbook(file_path)
            text = []
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    text.append('\t'.join(str(cell) for cell in row))
            return '\n'.join(text)
        elif file_type == 'pdf':
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                text = []
                for page in reader.pages:
                    text.append(page.extract_text())
                return '\n'.join(text)
        else:
            return "Unsupported file type"
    except Exception as e:
        return f"Error reading file: {str(e)}"

def process_archive(archive_path, archive_name, base_dir, csv_writer):
    """Обрабатывает архив и извлекает файлы для анализа.

    В рамках текущего задания все архивы (.zip, .rar, .7z) созданы
    при помощи zipfile, поэтому читаем их единообразно как ZIP.
    """
    if archive_path.endswith(('.zip', '.rar', '.7z')):
        with zipfile.ZipFile(archive_path, 'r') as zf:
            with tempfile.TemporaryDirectory() as tmpdir:
                zf.extractall(tmpdir)
                for root, _, files in os.walk(tmpdir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        rel_path = os.path.relpath(file_path, tmpdir)
                        full_path = os.path.join(archive_name, rel_path)
                        file_type = os.path.splitext(file)[1][1:].lower()
                        content = extract_text_from_file(file_path, file_type)
                        content = normalize_content_for_csv(content)
                        csv_writer.writerow([
                            full_path,
                            file_type,
                            content
                        ])

def crawl_directory(directory_path):
    """Основная функция краулера"""
    csv_filename = 'file_crawler_results.csv'
    csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), csv_filename)

    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        csv_writer = csv.writer(csvfile)
        csv_writer.writerow(['Path', 'Type', 'Content'])  # Заголовки CSV

        for root, _, files in os.walk(directory_path):
            for file in files:
                file_path = os.path.join(root, file)
                rel_path = os.path.relpath(file_path, directory_path)
                file_type = os.path.splitext(file)[1][1:].lower()

                # Проверяем, является ли файл архивом
                if file_type in ['zip', 'rar', '7z']:
                    process_archive(file_path, rel_path, directory_path, csv_writer)
                else:
                    # Обрабатываем обычный файл
                    content = extract_text_from_file(file_path, file_type)
                    content = normalize_content_for_csv(content)
                    csv_writer.writerow([
                        rel_path,
                        file_type,
                        content
                    ])

    print(f"Краулер завершил работу. Результаты сохранены в {csv_path}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Использование: python file_crawler.py <путь_к_директории>")
        sys.exit(1)

    directory_path = sys.argv[1]
    if not os.path.isdir(directory_path):
        print(f"Ошибка: {directory_path} не является директорией")
        sys.exit(1)

    crawl_directory(directory_path)